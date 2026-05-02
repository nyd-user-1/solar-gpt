"""Parse NYISO interconnection-queue xlsx snapshots into a single rows.json.

Run: python3 parse.py [optional/path/to/xlsx-dir]

Auto-discovers every *.xlsx in the directory next to this script (or the
directory passed as argv[1]). For each file it:
  1. Extracts snapshot_date from the filename (multiple regex patterns).
  2. Walks each sheet looking for a recognized header row (1-row or 2-row
     header). Maps sheets by name to status:
       'Active' / 'Interconnection Queue'  -> 'active'
       'Withdrawn'                          -> 'withdrawn'
       'In Service'                         -> 'in_service'
     Other sheets ('Cluster Projects', 'Affected System Studies', etc.)
     are skipped — extend SHEET_STATUS below if you want them.
  3. Fuzzy-matches column headers against KNOWN_COLUMNS and builds a
     column-name -> column-index map per file. No hardcoded indices.
  4. Emits one row per data row with keys matching the columns in
     solargpt.raw_nyiso_queue.

Per-file status is logged as: OK / SKIPPED / FAILED. The script never
crashes on a single bad file; the failing file is reported and parsing
continues. Total written count is reported at the end.

Output: rows.json next to this script.
"""
from __future__ import annotations
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
INPUT_DIR = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else HERE
OUT_PATH = HERE / "rows.json"

# ── Sheet -> status mapping ────────────────────────────────────────────────
# Compared case-insensitively, whitespace-stripped.
SHEET_STATUS = {
    "active": "active",
    "interconnection queue": "active",
    "withdrawn": "withdrawn",
    "in service": "in_service",
}

# ── Known columns ───────────────────────────────────────────────────────────
# For each canonical DB field, list header tokens that should match it. The
# matcher lowercases both sides, strips non-alphanumeric, and uses substring
# match. First successful match wins.
KNOWN_COLUMNS: dict[str, list[str]] = {
    "queue_pos": ["queuepos", "queueposition", "queue"],
    "developer": ["developer", "owner", "interconnectioncustomer", "ownerdeveloper"],
    "project_name": ["projectname"],
    "date_of_ir": ["dateofir"],
    "sp_mw": ["spmw", "summer"],
    "wp_mw": ["wpmw", "winter"],
    "type_fuel": ["typefuel"],
    "energy_storage_capability": ["energystoragecapability", "storagecapability"],
    "minimum_duration_full_discharge": ["minimumduration", "fulldischarge"],
    "county": ["county"],
    "state": ["state"],
    "zone": ["zone", "z"],
    "points_of_interconnection": ["pointsofinterconnection", "interconnectionpoint", "point"],
    "utility": ["utility"],
    "affected_transmission_owner": ["affectedtransmissionowner", "ato"],
    "s": ["s"],
    "last_updated_date": ["lastupdateddate", "lastupdate"],
    "availability_of_studies": ["availabilityofstudies"],
    "ia_tender_date": ["iatender", "iatenderdate"],
    "cy_fs_complete_date": ["cyfscomplete", "fscomplete", "sgiatender"],
    "proposed_in_service_date": ["proposedinservice", "initialbackfeed"],
    "proposed_sync_date": ["proposedsync", "initialsync"],
    "proposed_cod": ["proposedcod", "cod"],
}

# Column families that need date-ish parsing
DATE_FIELDS = {"date_of_ir", "last_updated_date", "ia_tender_date", "cy_fs_complete_date"}
NUMERIC_FIELDS = {"sp_mw", "wp_mw"}


# ── Filename -> snapshot_date ───────────────────────────────────────────────
DATE_PATTERNS = [
    # NYISO-Interconnection-Queue-MM-DD-YYYY.xlsx
    (re.compile(r"(\d{1,2})-(\d{1,2})-(\d{4})"), ("m", "d", "Y")),
    # NYISO-Interconnection-Queue-MM_DD_YYYY.xlsx
    (re.compile(r"(\d{1,2})_(\d{1,2})_(\d{4})"), ("m", "d", "Y")),
    # NYISO Interconnection Queue M.D.YY.xlsx
    (re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})"), ("m", "d", "Y")),
    # NYISO-Interconnection-Queue-MMDDYY.xlsx (053119 -> 2019-05-31)
    (re.compile(r"-(\d{2})(\d{2})(\d{2})\.xlsx$", re.I), ("m", "d", "y")),
    # NYISO-Interconnection-Queue-MMDDYYYY.xlsx
    (re.compile(r"-(\d{2})(\d{2})(\d{4})\.xlsx$", re.I), ("m", "d", "Y")),
    # YYYY-MM-DD anywhere in filename
    (re.compile(r"(\d{4})-(\d{1,2})-(\d{1,2})"), ("Y", "m", "d")),
]


def detect_snapshot_date(filename: str) -> date | None:
    for rx, parts in DATE_PATTERNS:
        m = rx.search(filename)
        if not m:
            continue
        try:
            vals = dict(zip(parts, m.groups()))
            year = int(vals["Y"]) if "Y" in vals else 2000 + int(vals["y"])
            if year < 100:
                year += 2000
            month = int(vals["m"])
            day = int(vals["d"])
            if not (1 <= month <= 12 and 1 <= day <= 31):
                continue
            return date(year, month, day)
        except (ValueError, KeyError):
            continue
    return None


# ── Header fuzzy matcher ────────────────────────────────────────────────────
def normalize_header(h) -> str:
    if h is None:
        return ""
    s = str(h).lower()
    return re.sub(r"[^a-z0-9]+", "", s)


def build_column_map(header_row: list, second_row: list | None = None) -> dict[str, int]:
    """Return {db_field: column_index} for every field we can identify."""
    # Combine row 1 + row 2 for split headers (NYISO 2019 format does this:
    # row 1 has 'Queue', row 2 has 'Pos.'; the joined header is 'queuepos').
    headers = []
    for i, h1 in enumerate(header_row):
        h2 = second_row[i] if second_row and i < len(second_row) else None
        combined = " ".join(str(x) for x in (h1, h2) if x is not None and str(x).strip())
        headers.append(normalize_header(combined))

    col_map: dict[str, int] = {}
    for field, tokens in KNOWN_COLUMNS.items():
        for idx, h in enumerate(headers):
            if not h:
                continue
            if any(t in h for t in tokens):
                # First match wins. Skip if column already used by another field.
                if idx not in col_map.values():
                    col_map[field] = idx
                    break
    return col_map


def looks_like_header(row: list) -> bool:
    """Heuristic: a row is probably a header if it has 'queue' OR 'pos' in any cell."""
    for cell in row:
        if cell is None:
            continue
        norm = normalize_header(cell)
        if "queue" in norm or "pos" in norm or "owner" in norm or "developer" in norm:
            return True
    return False


# ── Value coercion ──────────────────────────────────────────────────────────
def to_date(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        if not s or s.upper() in {"N/A", "NA", "TBD", "NONE", "I/S"}:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%m-%d-%Y", "%m-%d-%y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                pass
        return None
    return None


def to_num(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s or s.upper() in {"N/A", "NA", "TBD"}:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def to_text(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    s = str(v).strip()
    return s if s else None


def normalize_queue_pos(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return f"{int(v):04d}"
    s = str(v).strip()
    if not s:
        return None
    # Skip obvious non-data rows
    if s.upper() in {"NOTES:", "NOTES", "TOTAL", "TOTALS"}:
        return None
    if s.startswith("●") or s.startswith("*"):
        return None
    return s


# ── Per-file parser ─────────────────────────────────────────────────────────
def parse_file(path: Path) -> tuple[list[dict], dict]:
    """Returns (rows, status_dict). status_dict has counts per sheet."""
    snap = detect_snapshot_date(path.name)
    if snap is None:
        return [], {"status": "FAILED", "reason": "no date in filename"}

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return [], {"status": "FAILED", "reason": f"openpyxl: {e}"}

    rows = []
    sheet_counts = {}
    for sheet_name in wb.sheetnames:
        key = sheet_name.strip().lower()
        status = SHEET_STATUS.get(key)
        if status is None:
            continue
        try:
            ws = wb[sheet_name]
            sheet_rows, sheet_status = parse_sheet(ws, sheet_name, status, snap, path.name)
            rows.extend(sheet_rows)
            sheet_counts[sheet_name] = sheet_status
        except Exception as e:
            sheet_counts[sheet_name] = f"FAILED: {e}"

    if not rows:
        return [], {"status": "SKIPPED", "reason": "no recognized sheets or all empty", "sheets": sheet_counts}

    return rows, {"status": "OK", "rows": len(rows), "sheets": sheet_counts}


def parse_sheet(ws, sheet_name: str, status: str, snap: date, source_file: str):
    # Find the header row by scanning the first 5 rows.
    header_row_idx = None
    second_row_idx = None
    for r in range(1, 6):
        row = [c.value for c in ws[r]] if ws.max_row >= r else []
        if looks_like_header(row):
            header_row_idx = r
            # Check if next row is also a header continuation (for 2-row headers)
            next_row = [c.value for c in ws[r + 1]] if ws.max_row >= r + 1 else []
            if next_row and any(
                normalize_header(c) in {"pos", "owner", "developer", "projectname", "ofir"}
                for c in next_row if c is not None
            ):
                second_row_idx = r + 1
            break

    if header_row_idx is None:
        return [], "no header found"

    header_row = [c.value for c in ws[header_row_idx]]
    second_row = [c.value for c in ws[second_row_idx]] if second_row_idx else None
    col_map = build_column_map(header_row, second_row)

    if "queue_pos" not in col_map:
        return [], f"no queue_pos column (saw row {header_row_idx}: {header_row[:5]})"

    data_start = (second_row_idx or header_row_idx) + 1
    out = []
    for r in range(data_start, ws.max_row + 1):
        raw_row = [c.value for c in ws[r]]
        qp = normalize_queue_pos(raw_row[col_map["queue_pos"]] if col_map["queue_pos"] < len(raw_row) else None)
        if not qp:
            continue
        rec = {
            "queue_pos": qp,
            "snapshot_date": snap.isoformat(),
            "status": status,
            "source_file": source_file,
            "source_sheet": sheet_name,
        }
        for field, idx in col_map.items():
            if field == "queue_pos":
                continue
            val = raw_row[idx] if idx < len(raw_row) else None
            if field in DATE_FIELDS:
                rec[field] = to_date(val)
            elif field in NUMERIC_FIELDS:
                rec[field] = to_num(val)
            else:
                rec[field] = to_text(val)
        out.append(rec)

    return out, f"{len(out)} rows"


# ── Main ────────────────────────────────────────────────────────────────────
def main():
    files = sorted(INPUT_DIR.glob("*.xlsx"))
    if not files:
        print(f"No .xlsx files found in {INPUT_DIR}")
        sys.exit(1)

    print(f"Discovered {len(files)} xlsx file(s) in {INPUT_DIR}\n")

    all_rows: list[dict] = []
    summary = []
    for f in files:
        rows, status = parse_file(f)
        line = f"  {f.name:55} -> {status['status']}"
        if status["status"] == "OK":
            line += f"  ({status['rows']} rows from {len(status['sheets'])} sheet(s))"
        elif status["status"] in ("SKIPPED", "FAILED"):
            line += f"  [{status['reason']}]"
        print(line)
        all_rows.extend(rows)
        summary.append((f.name, status))

    # Dedup on (queue_pos, snapshot_date) — prefer active > in_service > withdrawn
    priority = {"active": 0, "in_service": 1, "withdrawn": 2}
    seen: dict[tuple[str, str], dict] = {}
    for r in all_rows:
        key = (r["queue_pos"], r["snapshot_date"])
        if key not in seen or priority.get(r["status"], 9) < priority.get(seen[key]["status"], 9):
            seen[key] = r
    deduped = list(seen.values())

    OUT_PATH.write_text(json.dumps(deduped, default=str))
    print(f"\nTotal: {len(all_rows)} parsed -> {len(deduped)} deduped -> wrote {OUT_PATH}")

    failed = [name for name, s in summary if s["status"] == "FAILED"]
    if failed:
        print(f"\n{len(failed)} file(s) failed — see status above. Re-run after fixing parser.")
        sys.exit(2)


if __name__ == "__main__":
    main()
